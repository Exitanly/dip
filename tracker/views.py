from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum, Avg, Count, Q
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login
from django.contrib import messages
from django.utils import timezone
from django.http import  JsonResponse, HttpResponse
from .models import Transaction, Category
from .forms import TransactionForm
from datetime import datetime, timedelta
import calendar
from calendar import month_name
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill


@login_required
def dashboard(request):
    user = request.user
    today = timezone.now().date()
    
       # ========== ОБРАБОТКА ФИЛЬТРА ==========
    # Получаем параметры из GET-запроса или из сессии
    period_type = request.GET.get('period_type')
    
    # Сбрасываем session принудительно для теста (временно)
    # request.session.flush()
    
    if period_type:
        request.session['period_type'] = period_type
        request.session['period_value'] = None
        request.session['start_date'] = None
        request.session['end_date'] = None
        
        if period_type == 'specific_month':
            month_val = request.GET.get('month_value')  
            if month_val:
                request.session['period_value'] = month_val
                print(f"[DEBUG] Сохранён месяц: {month_val}")
                
        elif period_type == 'year':
            year_val = request.GET.get('year_value')  
            if year_val:
                request.session['period_value'] = year_val
                print(f"[DEBUG] Сохранён год: {year_val}")
                
        elif period_type == 'custom':
            start_val = request.GET.get('start_date')
            end_val = request.GET.get('end_date')
            if start_val:
                request.session['start_date'] = start_val
            if end_val:
                request.session['end_date'] = end_val
            print(f"[DEBUG] Сохранён период: {start_val} - {end_val}")
    
    # Восстанавливаем из сессии
    period_type = request.session.get('period_type', 'current_month')
    
    start_date = None
    end_date = None
    period_display = ""
    
    # Определяем даты начала и конца периода
    if period_type == 'current_month':
        start_date = today.replace(day=1)
        end_date = today
        period_display = f"за {today.strftime('%B %Y')}"
        
    elif period_type == 'specific_month':
        month_str = request.session.get('period_value')
        print(f"[DEBUG] Загружен месяц из сессии: {month_str}")  # Отладка
        
        if month_str and '-' in month_str:
            try:
                year, month = map(int, month_str.split('-'))
                start_date = datetime(year, month, 1).date()
                if month == 12:
                    end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
                else:
                    end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
                period_display = f"за {month_name[month]} {year}"
                print(f"[DEBUG] Дата начала: {start_date}, конец: {end_date}")  # Отладка
            except (ValueError, TypeError) as e:
                print(f"[DEBUG] Ошибка парсинга: {e}")
                start_date = today.replace(day=1)
                end_date = today
                period_display = f"за {today.strftime('%B %Y')}"
        else:
            start_date = today.replace(day=1)
            end_date = today
            period_display = f"за {today.strftime('%B %Y')}"
            
    elif period_type == 'year':
        year_str = request.session.get('period_value')
        if year_str and year_str.isdigit():
            year = int(year_str)
            start_date = datetime(year, 1, 1).date()
            end_date = datetime(year, 12, 31).date()
            period_display = f"за {year} год"
        else:
            start_date = today.replace(day=1)
            end_date = today
            period_display = f"за {today.strftime('%B %Y')}"
            
    elif period_type == 'custom':
        start_date_str = request.session.get('start_date')
        end_date_str = request.session.get('end_date')
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                period_display = f"с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}"
            except (ValueError, TypeError):
                start_date = today.replace(day=1)
                end_date = today
                period_display = f"за {today.strftime('%B %Y')}"
        else:
            start_date = today.replace(day=1)
            end_date = today
            period_display = f"за {today.strftime('%B %Y')}"
    
    else:
        start_date = today.replace(day=1)
        end_date = today
        period_display = f"за {today.strftime('%B %Y')}"
    
    # Если по какой-то причине даты не установлены
    if not start_date or not end_date:
        start_date = today.replace(day=1)
        end_date = today
        period_display = f"за {today.strftime('%B %Y')}"
    
    # Все транзакции пользователя (без фильтра по дате для прогноза)
    all_transactions = Transaction.objects.filter(user=user)
    
    # Транзакции за выбранный период
    transactions = all_transactions.filter(date__gte=start_date, date__lte=end_date)
    
    # Доходы и расходы за период
    total_income = float(transactions.filter(type='income').aggregate(Sum('amount'))['amount__sum'] or 0)
    total_expense = float(transactions.filter(type='expense').aggregate(Sum('amount'))['amount__sum'] or 0)
    balance = total_income - total_expense
    
    # Расходы по категориям за период (для графика)
    expense_by_category = transactions.filter(type='expense').values('category__name').annotate(total=Sum('amount'))
    category_labels = [item['category__name'] for item in expense_by_category]
    category_data = [float(item['total']) for item in expense_by_category]
    
    # Расходы по категориям за период (для прогресс-баров)
    # Для прогресс-баров используем лимиты из категорий, но сравниваем с тратами за период
    monthly_expenses = transactions.filter(type='expense').values('category_id').annotate(spent=Sum('amount'))
    spent_dict = {item['category_id']: float(item['spent']) for item in monthly_expenses}
    
    # Категории с лимитами
    categories = Category.objects.filter(user=user, is_income=False)
    budget_data = []
    for category in categories:
        spent = spent_dict.get(category.id, 0)
        limit = float(category.budget_limit) if category.budget_limit else None
        percent = (spent / limit * 100) if limit and limit > 0 else None
        
        if limit is None:
            status = 'no_limit'
            color = 'secondary'
            alert = None
        elif percent < 70:
            status = 'good'
            color = 'success'
            alert = None
        elif percent < 100:
            status = 'warning'
            color = 'warning'
            alert = f'Осталось {limit - spent:.2f} ₽'
        else:
            status = 'danger'
            color = 'danger'
            alert = f'Перерасход на {spent - limit:.2f} ₽'
        
        budget_data.append({
            'id': category.id,
            'name': category.name,
            'spent': spent,
            'limit': limit,
            'percent': percent,
            'status': status,
            'color': color,
            'alert': alert,
        })
    
    # ========== АНАЛИТИКА (на основе всех данных) ==========
    
    # Прогноз на следующий месяц (на основе последних 3 месяцев)
    three_months_ago = today - timedelta(days=90)
    last_3_months_avg = all_transactions.filter(
        type='expense',
        date__gte=three_months_ago,
        date__lte=today
    ).aggregate(avg_monthly=Avg('amount'))['avg_monthly']
    
    if last_3_months_avg:
        predicted_expenses = float(last_3_months_avg) * 1.05
    else:
        predicted_expenses = 0
    
    # Топ-3 категории для сокращения (на основе текущего периода)
    top_categories_to_reduce = []
    for budget in budget_data:
        if budget['limit'] and budget['percent'] and budget['percent'] > 70:
            spent = float(budget['spent'])
            limit = float(budget['limit'])
            potential_saving = spent - (limit * 0.7)
            if potential_saving > 0:
                top_categories_to_reduce.append({
                    'name': budget['name'],
                    'spent': spent,
                    'limit': limit,
                    'percent': round(budget['percent'], 1),
                    'potential_saving': round(potential_saving, 2),
                })
    top_categories_to_reduce = sorted(top_categories_to_reduce, key=lambda x: x['potential_saving'], reverse=True)[:3]
    
    # Умные советы
    tips = []
    
    if total_expense > total_income:
        tips.append({
            'icon': '⚠️',
            'title': 'Расходы превышают доходы',
            'message': f'Ваши расходы ({total_expense:.0f} ₽) превышают доходы ({total_income:.0f} ₽) на {total_expense - total_income:.0f} ₽.',
            'type': 'danger'
        })
    
    if total_income > 0 and balance < (total_income * 0.05):
        tips.append({
            'icon': '🏦',
            'title': 'Низкая финансовая подушка',
            'message': 'Старайтесь откладывать минимум 5-10% от доходов.',
            'type': 'warning'
        })
    
    over_budget_categories = [b for b in budget_data if b.get('status') == 'danger']
    if over_budget_categories:
        names = ', '.join([b['name'] for b in over_budget_categories[:2]])
        tips.append({
            'icon': '📈',
            'title': 'Перерасход по бюджету',
            'message': f'Вы превысили лимит в категориях: {names}.',
            'type': 'warning'
        })
    
    # Динамика расходов по дням за последние 30 дней (всегда за последние 30 дней)
    thirty_days_ago = today - timedelta(days=30)
    daily_expenses = all_transactions.filter(
        type='expense',
        date__gte=thirty_days_ago,
        date__lte=today
    ).values('date').annotate(total=Sum('amount')).order_by('date')
    
    daily_labels = []
    daily_data = []
    for item in daily_expenses:
        date_obj = item['date']
        if isinstance(date_obj, str):
            date_obj = datetime.strptime(date_obj, '%Y-%m-%d').date()
        daily_labels.append(date_obj.strftime('%d.%m'))
        daily_data.append(float(item['total']))
    
    # Последние 5 транзакций за период
    recent_transactions = transactions[:5]
    
    # Список доступных месяцев для выпадающего списка
    available_months = []
    dates = all_transactions.dates('date', 'month', order='DESC')
    for dt in dates:
        available_months.append(dt.strftime('%Y-%m'))
    
    context = {
        'total_income': total_income,
        'total_expense': total_expense,
        'balance': balance,
        'category_labels': category_labels,
        'category_data': category_data,
        'budget_data': budget_data,
        'recent_transactions': recent_transactions,
        'period_display': period_display,
        'period_type': period_type,
        'available_months': available_months,
        'selected_month': request.session.get('period_value', today.strftime('%Y-%m')),
        'selected_year': request.session.get('period_value', str(today.year)),
        'start_date_val': request.session.get('start_date', today.replace(day=1).strftime('%Y-%m-%d')),
        'end_date_val': request.session.get('end_date', today.strftime('%Y-%m-%d')),
        # Аналитика
        'predicted_expenses': round(predicted_expenses, 2),
        'top_categories_to_reduce': top_categories_to_reduce,
        'tips': tips,
        'daily_labels': daily_labels,
        'daily_data': daily_data,
    }
    
    return render(request, 'dashboard.html', context)

@login_required
def add_transaction(request):
    if request.method == 'POST':
        # Получаем данные из POST-запроса
        transaction_type = request.POST.get('type')
        category_id = request.POST.get('category')
        amount = request.POST.get('amount')
        date = request.POST.get('date')
        description = request.POST.get('description')
        
        # Создаём транзакцию
        transaction = Transaction(
            user=request.user,
            type=transaction_type,
            category_id=category_id,
            amount=amount,
            date=date,
            description=description
        )
        transaction.save()
        return redirect('dashboard')
    else:
        expense_categories = Category.objects.filter(user=request.user, is_income=False).values('id', 'name', 'budget_limit')
        income_categories = Category.objects.filter(user=request.user, is_income=True).values('id', 'name', 'budget_limit')
        
        context = {
            'title': 'Добавить операцию',
            'expense_categories': expense_categories,
            'income_categories': income_categories,
        }
        
        # Если есть параметр type в GET (например, при редактировании), передаём его
        if 'type' in request.GET:
            context['selected_type'] = request.GET.get('type')
        
        return render(request, 'transaction_form.html', context)

@login_required
def transaction_list(request):
    transactions = Transaction.objects.filter(user=request.user)
    
    # Фильтрация
    transaction_type = request.GET.get('type')
    category_id = request.GET.get('category')
    
    if transaction_type:
        transactions = transactions.filter(type=transaction_type)
    if category_id:
        transactions = transactions.filter(category_id=category_id)
    
    categories = Category.objects.filter(user=request.user)
    
    return render(request, 'transaction_list.html', {
        'transactions': transactions,
        'categories': categories,
        'current_type': transaction_type,
        'current_category': category_id,
    })

@login_required
def edit_transaction(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk, user=request.user)
    
    if request.method == 'POST':
        # Обновляем транзакцию
        transaction.type = request.POST.get('type')
        transaction.category_id = request.POST.get('category')
        transaction.amount = request.POST.get('amount')
        transaction.date = request.POST.get('date')
        transaction.description = request.POST.get('description')
        transaction.save()
        return redirect('transaction_list')
    else:
        # Разделяем категории
        expense_categories = Category.objects.filter(user=request.user, is_income=False)
        income_categories = Category.objects.filter(user=request.user, is_income=True)
        
        context = {
            'title': 'Редактировать операцию',
            'expense_categories': expense_categories,
            'income_categories': income_categories,
            'form': transaction,  # передаём транзакцию для отображения текущих значений
        }
        
        return render(request, 'transaction_form.html', context)

@login_required
def delete_transaction(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk, user=request.user)
    if request.method == 'POST':
        transaction.delete()
        return redirect('transaction_list')
    
    return render(request, 'transaction_confirm_delete.html', {'transaction': transaction})

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            
            # Создаём стандартные категории для нового пользователя
            default_categories = [
                {'name': 'Еда', 'is_income': False},
                {'name': 'Транспорт', 'is_income': False},
                {'name': 'Кафе и рестораны', 'is_income': False},
                {'name': 'Супермаркеты', 'is_income': False},
                {'name': 'Связь и интернет', 'is_income': False},
                {'name': 'Зарплата', 'is_income': True},
                {'name': 'Подработка', 'is_income': True},
            ]
            
            for cat_data in default_categories:
                Category.objects.create(
                    user=user,
                    name=cat_data['name'],
                    is_income=cat_data['is_income'],
                    budget_limit=None  # можно потом установить
                )
            
            login(request, user)
            return redirect('dashboard')
    else:
        form = UserCreationForm()
    
    return render(request, 'registration/register.html', {'form': form})

@login_required
def category_list(request):
    categories = Category.objects.filter(user=request.user)
    return render(request, 'category_list.html', {'categories': categories})

@login_required
def category_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        is_income = request.POST.get('is_income') == 'on'
        budget_limit = request.POST.get('budget_limit') or None
        
        Category.objects.create(
            user=request.user,
            name=name,
            is_income=is_income,
            budget_limit=budget_limit
        )
        messages.success(request, f'Категория "{name}" создана!')
        return redirect('category_list')
    
    return render(request, 'category_form.html', {'title': 'Создать категорию'})

@login_required
def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk, user=request.user)
    
    if request.method == 'POST':
        category.name = request.POST.get('name')
        category.is_income = request.POST.get('is_income') == 'on'
        category.budget_limit = request.POST.get('budget_limit') or None
        category.save()
        messages.success(request, f'Категория "{category.name}" обновлена!')
        return redirect('category_list')
    
    return render(request, 'category_form.html', {'category': category, 'title': 'Редактировать категорию'})

@login_required
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk, user=request.user)
    if request.method == 'POST':
        category_name = category.name
        category.delete()
        messages.success(request, f'Категория "{category_name}" удалена!')
        return redirect('category_list')
    
    return render(request, 'category_confirm_delete.html', {'category': category})

@login_required
def category_spent_api(request, category_id):
    """API для получения суммы расходов по категории за текущий месяц"""
    user = request.user
    today = timezone.now().date()
    first_day_of_month = today.replace(day=1)
    
    # Проверяем, что категория принадлежит пользователю
    try:
        category = Category.objects.get(id=category_id, user=user)
    except Category.DoesNotExist:
        return JsonResponse({'error': 'Категория не найдена'}, status=404)
    
    # Сумма расходов за текущий месяц
    spent = Transaction.objects.filter(
        user=user,
        category=category,
        type='expense',
        date__gte=first_day_of_month,
        date__lte=today
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    return JsonResponse({
        'category_id': category_id,
        'spent': float(spent),
        'limit': float(category.budget_limit) if category.budget_limit else None,
    })

@login_required
def export_to_excel(request):
    """Экспорт данных за текущий месяц в Excel с диаграммой"""
    user = request.user
    today = datetime.now().date()
    first_day = today.replace(day=1)
    last_day = (first_day + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    # Получаем операции за месяц
    transactions = Transaction.objects.filter(
        user=user,
        date__gte=first_day,
        date__lte=last_day
    ).select_related('category')
    
    # Получаем категории с расходами за месяц
    expense_categories = Category.objects.filter(user=user, is_income=False)
    
    # Считаем потраченное по категориям
    category_spending = {}
    for cat in expense_categories:
        spent = transactions.filter(category=cat, type='expense').aggregate(Sum('amount'))['amount__sum'] or 0
        category_spending[cat.id] = {
            'name': cat.name,
            'spent': float(spent),
            'limit': float(cat.budget_limit) if cat.budget_limit else None,
        }
    
    # Создаём Excel-файл
    wb = openpyxl.Workbook()
    
    # --- Лист 1: Все операции ---
    ws_operations = wb.active
    ws_operations.title = "Операции"
    
    # Заголовки
    headers = ['Дата', 'Тип', 'Категория', 'Сумма', 'Описание']
    for col, header in enumerate(headers, 1):
        cell = ws_operations.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Данные
    for row, t in enumerate(transactions, 2):
        ws_operations.cell(row=row, column=1, value=t.date.strftime('%d.%m.%Y'))
        ws_operations.cell(row=row, column=2, value='Расход' if t.type == 'expense' else 'Доход')
        ws_operations.cell(row=row, column=3, value=t.category.name if t.category else '-')
        ws_operations.cell(row=row, column=4, value=float(t.amount))
        ws_operations.cell(row=row, column=5, value=t.description or '')
    
    # --- Лист 2: Сводка по категориям ---
    ws_summary = wb.create_sheet("Сводка по бюджету")
    
    summary_headers = ['Категория', 'Потрачено (₽)', 'Лимит (₽)', 'Остаток (₽)', 'Выполнение (%)']
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    row = 2
    for cat_data in category_spending.values():
        spent = cat_data['spent']
        limit = cat_data['limit']
        
        ws_summary.cell(row=row, column=1, value=cat_data['name'])
        ws_summary.cell(row=row, column=2, value=spent)
        
        if limit:
            remaining = limit - spent
            percent = (spent / limit * 100) if limit > 0 else 0
            ws_summary.cell(row=row, column=3, value=limit)
            ws_summary.cell(row=row, column=4, value=round(remaining, 2))
            ws_summary.cell(row=row, column=5, value=round(percent, 2))
            
            # Цвет ячейки в зависимости от процента
            if percent > 100:
                ws_summary.cell(row=row, column=5).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        else:
            ws_summary.cell(row=row, column=3, value='Нет лимита')
            ws_summary.cell(row=row, column=4, value='-')
            ws_summary.cell(row=row, column=5, value='-')
        
        row += 1
    
    # --- Лист 3: Диаграмма расходов (Excel-график) ---
    ws_chart = wb.create_sheet("График расходов")
    
    # Данные для диаграммы
    chart_data = [[cat_data['name'], cat_data['spent']] for cat_data in category_spending.values() if cat_data['spent'] > 0]
    
    if chart_data:
        ws_chart.cell(row=1, column=1, value="Категория")
        ws_chart.cell(row=1, column=2, value="Сумма (₽)")
        
        for i, (name, spent) in enumerate(chart_data, 2):
            ws_chart.cell(row=i, column=1, value=name)
            ws_chart.cell(row=i, column=2, value=spent)
        
        # Создаём круговую диаграмму
        pie = PieChart()
        data = Reference(ws_chart, min_col=2, min_row=1, max_row=len(chart_data) + 1)
        labels = Reference(ws_chart, min_col=1, min_row=2, max_row=len(chart_data) + 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Структура расходов"
        pie.height = 15
        pie.width = 20
        
        ws_chart.add_chart(pie, "D1")
    
    # Настройка ширины колонок
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_operations.column_dimensions[col].width = 18
        ws_summary.column_dimensions[col].width = 18
    
    # Формируем ответ
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="finance_report_{first_day.strftime("%Y_%m")}.xlsx"'
    wb.save(response)
    
    return response
